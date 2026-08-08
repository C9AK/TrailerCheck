"""Manager data export — daily pickups as an Excel workbook, one sheet per
calendar day in the requested range (sheet name = that day's date), for
Excel / Google Sheets.

All relational fields are resolved to human-readable strings (MC name,
employee username, approving QC username, flag category labels) — no UUIDs.
"""

import io
from collections import defaultdict
from datetime import date as date_type
from datetime import datetime, time, timedelta, timezone

from fastapi import APIRouter, Depends, HTTPException, Query, Response, status
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.orm import Session, joinedload, selectinload

from app.api.deps import require_roles
from app.core.database import get_db
from app.models import (
    AuditEvent,
    AuditLog,
    ErrorCategory,
    PickupTicket,
    QCAuditFlag,
    User,
    UserRole,
)
from app.models.enums import KPRA_GROUP_LABELS

router = APIRouter(tags=["export"])

CATEGORY_LABELS: dict[ErrorCategory, str] = {
    ErrorCategory.Missing_Inspection: "Missing inspection",
    ErrorCategory.Missing_Sticker: "Missing sticker",
    ErrorCategory.Missing_Registration: "Missing registration",
    ErrorCategory.Missed_KPRA_Reminder: "Didn't remind the driver about KPRA law",
    ErrorCategory.PTI_Video_Missing_Light_Test: "PTI video wasn't with the light test",
    ErrorCategory.Didnt_Text_In_Group: "Didn't text in the group",
    ErrorCategory.Missing_BOL: "Missing BOL",
    ErrorCategory.Incorrect_Weight: "Incorrect weight",
    ErrorCategory.Missed_PTI: "Missed PTI",
    ErrorCategory.Other: "Other",
}

HEADERS = [
    # R40: fixed report order the manager asked for —
    "Truck #", "Notes", "Checked", "Insp", "Reg", "Sticker", "BOL", "Weight",
    "TRL COND", "Scale", "Name", "LOT", "Date of Check", "Provider", "Status", "CA/FL",
    # ...then everything else, appended as extra context.
    "Created At (UTC)", "Created By", "Truck Model", "Location", "Fuel %",
    "KPRA Destination Group", "Needs Scale", "Flag Categories", "Flag Notes",
    "Flagged By", "Approved By", "Approved At (UTC)",
]

# One sheet is created per calendar day in the requested range (even empty
# ones) — cap the span so a fat-fingered range doesn't spend minutes
# generating thousands of blank sheets.
MAX_EXPORT_DAYS = 366

HEADER_FONT = Font(bold=True)


def _yn(value: bool) -> str:
    return "Yes" if value else "No"


def _fmt_dt(dt: datetime | None) -> str:
    if dt is None:
        return ""
    return dt.strftime("%Y-%m-%d %H:%M:%S")


def _fmt_date(dt: datetime | None) -> str:
    if dt is None:
        return ""
    return dt.strftime("%Y-%m-%d")


def _write_sheet(ws: Worksheet, tickets: list[PickupTicket], approvals: dict) -> None:
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = HEADER_FONT
    ws.freeze_panes = "A2"

    for t in tickets:
        categories = "; ".join(
            dict.fromkeys(CATEGORY_LABELS[f.error_category] for f in t.audit_flags)
        )
        flag_notes = "; ".join(
            dict.fromkeys(f.notes.strip() for f in t.audit_flags if f.notes and f.notes.strip())
        )
        flaggers = "; ".join(dict.fromkeys(f.flagger.username for f in t.audit_flags))
        approved_by, approved_at = approvals.get(t.id, ("", None))
        scale = _yn(t.scale_ticket_received) if t.needs_scale else "N/A"

        ws.append([
            # R40: fixed report order —
            t.truck_number,
            t.condition_notes or "",
            _yn(t.pti_verified),
            _yn(t.inspection_paper_verified),
            _yn(t.registration_verified),
            _yn(t.sticker_verified),
            _yn(t.bol_present),
            t.weight or "",
            t.trailer_condition.value if t.trailer_condition else "",
            scale,
            t.driver_name or "",
            _yn(t.is_lot_trailer),
            _fmt_date(t.created_at),
            t.motor_carrier.name,
            t.state.value,
            _yn(t.kpra_group == "CA_FL_40FT"),
            # ...then everything else.
            _fmt_dt(t.created_at),
            t.creator.username,
            t.truck_model or "",
            t.truck_location or "",
            f"{t.fuel_percentage:.0f}" if t.fuel_percentage is not None else "",
            KPRA_GROUP_LABELS.get(t.kpra_group, "") if t.kpra_group else "",
            _yn(t.needs_scale),
            categories,
            flag_notes,
            flaggers,
            approved_by,
            _fmt_dt(approved_at),
        ])

    for col_cells in ws.columns:
        width = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(width + 2, 10), 40)


@router.get("/api/export/pickups")
def export_pickups(
    start_date: date_type = Query(..., description="First day to export (YYYY-MM-DD)"),
    end_date: date_type | None = Query(
        None, description="Last day to export, inclusive (YYYY-MM-DD) — defaults to start_date"
    ),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(UserRole.manager)),
):
    # R40: was a single `date` param (daily-only) — the Archive UI's "To"
    # date was silently ignored, so a multi-day export always came back as
    # just the first day. end_date defaults to start_date for a same-day export.
    range_end = end_date or start_date
    if range_end < start_date:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="end_date cannot be before start_date.",
        )
    if (range_end - start_date).days + 1 > MAX_EXPORT_DAYS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Range too large for a per-day export — max {MAX_EXPORT_DAYS} days.",
        )
    range_start = datetime.combine(start_date, time.min, timezone.utc)
    range_end_dt = datetime.combine(range_end, time.max, timezone.utc)

    tickets = (
        db.scalars(
            select(PickupTicket)
            .options(
                joinedload(PickupTicket.creator),
                joinedload(PickupTicket.motor_carrier),
                selectinload(PickupTicket.audit_flags).joinedload(QCAuditFlag.flagger),
            )
            .where(PickupTicket.created_at >= range_start, PickupTicket.created_at <= range_end_dt)
            .order_by(PickupTicket.created_at.asc())
        )
        .unique()
        .all()
    )

    # Latest approval per ticket -> (QC username, timestamp)
    approvals: dict = {}
    if tickets:
        rows = db.execute(
            select(AuditLog.ticket_id, AuditLog.created_at, User.username)
            .join(User, AuditLog.actor_id == User.id)
            .where(
                AuditLog.event == AuditEvent.TICKET_APPROVED,
                AuditLog.ticket_id.in_([t.id for t in tickets]),
            )
            .order_by(AuditLog.created_at.asc())
        ).all()
        for ticket_id, created_at, username in rows:
            approvals[ticket_id] = (username, created_at)

    # Bucket tickets by their UTC calendar day — the same boundary used
    # above to pull the range in the first place — so a ticket always lands
    # on the sheet matching the day it was queried under.
    by_day: dict[date_type, list[PickupTicket]] = defaultdict(list)
    for t in tickets:
        by_day[t.created_at.date()].append(t)

    wb = Workbook()
    wb.remove(wb.active)  # replaced by one real sheet per day below

    # One sheet per calendar day in the range, in order — even days with no
    # pickups get a (header-only) sheet, so a quiet day still shows up
    # rather than silently vanishing from the export.
    day = start_date
    while day <= range_end:
        ws = wb.create_sheet(title=day.isoformat())
        _write_sheet(ws, by_day.get(day, []), approvals)
        day += timedelta(days=1)

    buffer = io.BytesIO()
    wb.save(buffer)
    filename = (
        f"pickups_{start_date.isoformat()}.xlsx"
        if range_end == start_date
        else f"pickups_{start_date.isoformat()}_to_{range_end.isoformat()}.xlsx"
    )
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
